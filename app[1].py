# app.py — Gestor de Facturas · Autónomos España
# Ejecutar con: streamlit run app.py

import streamlit as st
import pdfplumber
import pandas as pd
import io
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from dateutil import parser as dateparser

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER

CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Serif+Display&display=swap');
:root {
    --bg:        #F7F5F2;
    --surface:   #FFFFFF;
    --border:    #E4E0DB;
    --text-1:    #1C1917;
    --text-2:    #6B6460;
    --accent:    #2D6A4F;
    --accent-lt: #D8EFE4;
    --amber:     #B45309;
    --amber-lt:  #FEF3C7;
    --red-lt:    #FEE2E2;
    --shadow:    0 1px 3px rgba(0,0,0,.06), 0 4px 16px rgba(0,0,0,.04);
}
html, body, [data-testid="stAppViewContainer"] {
    background: var(--bg) !important;
    font-family: 'DM Sans', sans-serif !important;
    color: var(--text-1) !important;
}
[data-testid="stHeader"] { background: transparent !important; }
[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid var(--border) !important;
}
[data-testid="stSidebar"] * { font-family: 'DM Sans', sans-serif !important; }
.main-title {
    font-family: 'DM Serif Display', serif !important;
    font-size: 2.1rem !important;
    font-weight: 400 !important;
    color: var(--text-1) !important;
    letter-spacing: -.5px;
    margin: 0 0 .2rem 0;
}
.main-subtitle { font-size: .9rem; color: var(--text-2); margin: 0 0 1.2rem 0; }
.metric-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 1.1rem 1.3rem;
    box-shadow: var(--shadow);
    text-align: center;
}
.metric-label {
    font-size: .72rem; font-weight: 500; text-transform: uppercase;
    letter-spacing: .06em; color: var(--text-2); margin-bottom: .35rem;
}
.metric-value { font-size: 1.55rem; font-weight: 600; color: var(--accent); letter-spacing: -.5px; }
.metric-value.neutral { color: var(--text-1); }
.section-card {
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 14px; padding: 1.5rem 1.6rem;
    box-shadow: var(--shadow); margin-bottom: 1.2rem;
}
.section-title {
    font-size: .78rem; font-weight: 600; text-transform: uppercase;
    letter-spacing: .1em; color: var(--text-2);
    margin: 0 0 1rem 0; padding-bottom: .6rem; border-bottom: 1px solid var(--border);
}
[data-testid="stFileUploader"] {
    background: var(--surface) !important;
    border: 1.5px dashed var(--border) !important;
    border-radius: 12px !important;
}
[data-testid="stButton"] > button[kind="primary"] {
    background: var(--accent) !important; color: white !important;
    border: none !important; border-radius: 8px !important;
    font-weight: 500 !important; font-family: 'DM Sans', sans-serif !important;
    transition: opacity .15s ease;
}
[data-testid="stButton"] > button[kind="primary"]:hover { opacity: .88 !important; }
[data-testid="stDownloadButton"] > button {
    background: var(--surface) !important; color: var(--text-1) !important;
    border: 1px solid var(--border) !important; border-radius: 8px !important;
    font-family: 'DM Sans', sans-serif !important; font-weight: 500 !important;
    transition: border-color .15s, box-shadow .15s;
}
[data-testid="stDownloadButton"] > button:hover {
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 2px var(--accent-lt) !important;
}
[data-testid="stDataEditor"] {
    border-radius: 10px !important; overflow: hidden !important;
    border: 1px solid var(--border) !important;
}
.alert-warn {
    background: var(--amber-lt); border: 1px solid #FCD34D; border-radius: 8px;
    padding: .65rem 1rem; font-size: .84rem; color: var(--amber); margin-bottom: .8rem;
}
.alert-err {
    background: var(--red-lt); border: 1px solid #FCA5A5; border-radius: 8px;
    padding: .65rem 1rem; font-size: .84rem; color: #B91C1C; margin-bottom: .8rem;
}
.alert-ok {
    background: var(--accent-lt); border: 1px solid #86EFAC; border-radius: 8px;
    padding: .65rem 1rem; font-size: .84rem; color: var(--accent); margin-bottom: .8rem;
}
.alert-priv {
    background: #F1F5F9; border: 1px solid #CBD5E1; border-radius: 8px;
    padding: .65rem 1rem; font-size: .82rem; color: #475569; margin-bottom: 1.2rem;
}
.sidebar-logo { font-family: 'DM Serif Display', serif; font-size: 1.3rem; color: var(--accent); margin-bottom: .3rem; }
.sidebar-tagline { font-size: .78rem; color: var(--text-2); margin-bottom: 1.2rem; line-height: 1.4; }
.sidebar-label {
    font-size: .72rem; font-weight: 600; text-transform: uppercase;
    letter-spacing: .08em; color: var(--text-2); margin: 1rem 0 .3rem 0;
}
#MainMenu, footer { visibility: hidden; }
</style>
"""

st.set_page_config(page_title="Gestor de Facturas · España", page_icon="🧾", layout="wide")
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ── Columnas y constantes ──────────────────────────────────────────────────────

CONCEPTO_OPTIONS = [
    'Seguros', 'Aplicaciones / Software', 'Materiales', 'Servicios profesionales',
    'Transporte', 'Alojamiento', 'Publicidad / Marketing', 'Telecomunicaciones',
    'Suministros', 'Formación', 'Gestoría / Asesoría', 'Arrendamiento',
    'Equipos / Hardware', 'Otros gastos',
]
TIPO_DOC_OPTIONS = ['Factura', 'Factura simplificada', 'Ticket', 'Recibo', 'Nota de cargo', 'Otro']
COLUMNS = [
    'tipo_documento', 'fecha', 'numero_factura', 'nombre_proveedor', 'nif',
    'concepto', 'direccion', 'cp_pais', 'base_imponible',
    'tipo_iva', 'porcion_iva', 'tipo_retencion', 'porcion_retencion',
    'total', 'validacion', 'error',
]
DISPLAY_NAMES = {
    'tipo_documento':   'Tipo',
    'fecha':            'Fecha',
    'numero_factura':   'Nº Factura',
    'nombre_proveedor': 'Nombre Proveedor',
    'nif':              'NIF/CIF',
    'concepto':         'Concepto',
    'direccion':        'Dirección',
    'cp_pais':          'CP / País',
    'base_imponible':   'Base Imponible (Mod.303)',
    'tipo_iva':         'Tipo IVA (%)',
    'porcion_iva':      'Porción IVA (€)',
    'tipo_retencion':   'Tipo Retención (%)',
    'porcion_retencion':'Porción Retención (€)',
    'total':            'Total (€)',
    'validacion':       'Validación',
    'error':            'Error',
}
CONCEPTO_KEYWORDS = {
    'Seguros':                  ['seguro','prima','cobertura','axa','mapfre','allianz','zurich'],
    'Aplicaciones / Software':  ['software','saas','licencia','subscri','adobe','microsoft','google workspace',
                                  'slack','notion','dropbox','github','hosting','dominio','cloud','aws','azure'],
    'Materiales':               ['material','papelería','tóner','cartucho','consumible','papel','tinta','ferretería'],
    'Telecomunicaciones':       ['telefon','móvil','internet','fibra','movistar','vodafone','orange','yoigo','digi'],
    'Publicidad / Marketing':   ['publicidad','marketing','anuncio','ads','diseño','impresión'],
    'Transporte':               ['taxi','uber','cabify','renfe','avión','vuelo','transporte','gasolina',
                                  'combustible','peaje','parking'],
    'Alojamiento':              ['hotel','hostal','airbnb','alojamiento'],
    'Formación':                ['formación','curso','udemy','masterclass','training','seminario'],
    'Gestoría / Asesoría':      ['gestoría','asesoría','gestor','asesor','notaría','abogado','legal','contabilidad'],
    'Arrendamiento':            ['alquiler','arrend','renta','local','oficina','coworking'],
    'Suministros':              ['electricidad','agua','gas','energía','iberdrola','endesa','naturgy'],
    'Equipos / Hardware':       ['ordenador','portátil','monitor','teclado','impresora','disco','tablet','hardware'],
}

def guess_concepto(text: str) -> str:
    t = text.lower()
    for concepto, kws in CONCEPTO_KEYWORDS.items():
        if any(k in t for k in kws):
            return concepto
    return 'Otros gastos'

# ── Utilidades ─────────────────────────────────────────────────────────────────

def parse_amount(text):
    if not text:
        return None
    text = str(text).strip()
    if re.search(r'\d\.\d{3},', text):
        text = text.replace('.','').replace(',','.')
    elif ',' in text and '.' not in text:
        text = text.replace(',','.')
    elif ',' in text and '.' in text:
        if text.rfind(',') > text.rfind('.'):
            text = text.replace('.','').replace(',','.')
        else:
            text = text.replace(',','')
    text = re.sub(r'[^\d.\-]', '', text)
    try:
        return float(text)
    except ValueError:
        return None

def parse_date(text: str) -> str:
    if not text:
        return '-'
    try:
        dt = dateparser.parse(text, dayfirst=True)
        return dt.strftime('%d/%m/%Y') if dt else '-'
    except Exception:
        return '-'

def safe_float(val) -> float:
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0

def fmt_eur(val) -> str:
    v = safe_float(val)
    if v == 0:
        return '-'
    s = f"{v:,.2f}".replace(',','X').replace('.',',').replace('X','.')
    return f"{s} €"

def validate_invoice(row: dict) -> str:
    base  = safe_float(row.get('base_imponible'))
    cuota = safe_float(row.get('porcion_iva'))
    ret   = safe_float(row.get('porcion_retencion'))
    total = safe_float(row.get('total'))
    if total == 0:
        return '-'
    return '✓ OK' if abs(base + cuota - ret - total) <= 0.03 else '⚠ Revisar'

def _empty_row(filename: str) -> dict:
    row = {c: '-' for c in COLUMNS}
    row['error'] = ''
    return row

# ── Patrones regex ─────────────────────────────────────────────────────────────

PATTERNS = {
    'fecha': [
        r'fecha[^:\n]*[:\s]+(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
        r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{4})',
        r'(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})',
        r'date[^:\n]*[:\s]+(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
    ],
    'numero': [
        r'n[uú]mero\s*(?:de\s*)?factura[^:\n]*[:\s]+([A-Z0-9\-/]+)',
        r'factura\s*n[uú]m\.?[^:\n]*[:\s]+([A-Z0-9\-/]+)',
        r'n\.?[oº]?\s*factura[^:\n]*[:\s]+([A-Z0-9\-/]+)',
        r'invoice\s*(?:n[o.]?|number)[^:\n]*[:\s]+([A-Z0-9\-/]+)',
        r'^\s*([A-Z]{1,3}[\-/]?\d{4}[\-/]\d{2,6})\s*$',
    ],
    'nif': [
        r'(?:nif|cif|vat|n\.i\.f|c\.i\.f)[:\s.]+([A-Z0-9]{7,12})',
        r'\b([A-Z]\d{7}[A-Z0-9])\b',
        r'\b(\d{8}[A-Z])\b',
    ],
    'nombre': [
        r'(?:raz[oó]n\s*social|empresa)[^:\n]*[:\s]+([^\n]{3,70})',
        r'(?:emisor|proveedor|vendedor|facturado\s*por)[^:\n]*[:\s]+([^\n]{3,70})',
    ],
    'direccion': [
        r'(?:direcci[oó]n|domicilio|address)[^:\n]*[:\s]+([^\n]{5,80})',
        r'(?:calle|avda?\.?|plaza|paseo)\s+([^\n]{5,70})',
        r'(?:c/|cl\.)\s*([^\n]{5,70})',
    ],
    'base': [
        r'base\s*imponible[^:\n€$]*[:\s€$]+([\d.,]+)',
        r'base[^:\n€$]*[:\s€$]+([\d.,]+)',
        r'subtotal[^:\n€$]*[:\s€$]+([\d.,]+)',
    ],
    'pct_iva': [
        r'i\.?v\.?a\.?\s*[\(%]?\s*([\d,\.]+)\s*[\(%]',
        r'([\d,\.]+)\s*%\s*i\.?v\.?a',
        r'tipo\s*(?:de\s*)?i\.?v\.?a[^:\n]*[:\s]+([\d,\.]+)',
    ],
    'cuota_iva': [
        r'cuota\s*i\.?v\.?a[^:\n€$]*[:\s€$]+([\d.,]+)',
        r'i\.?v\.?a[^:\n€$%]*[:\s€$]+([\d.,]+)',
        r'importe\s*i\.?v\.?a[^:\n€$]*[:\s€$]+([\d.,]+)',
    ],
    'pct_ret': [
        r'retenci[oó]n\s*[\(%]?\s*([\d,\.]+)\s*[\(%]',
        r'irpf\s*[\(%]?\s*([\d,\.]+)\s*[\(%]',
        r'([\d,\.]+)\s*%\s*(?:retenci[oó]n|irpf)',
    ],
    'imp_ret': [
        r'retenci[oó]n[^:\n€$%]*[:\s€$]+([\d.,]+)',
        r'irpf[^:\n€$%]*[:\s€$]+([\d.,]+)',
    ],
    'total': [
        r'total\s*(?:a\s*pagar|factura|general|neto)?[^:\n€$]*[:\s€$]+([\d.,]+)',
        r'importe\s*total[^:\n€$]*[:\s€$]+([\d.,]+)',
        r'grand\s*total[^:\n€$]*[:\s€$]+([\d.,]+)',
    ],
}

def extract_pat(text: str, key: str) -> str:
    for pat in PATTERNS.get(key, []):
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            return m.group(1).strip()
    return ''

def detect_tipo_doc(text: str) -> str:
    t = text.lower()
    if re.search(r'factura\s*simplificada', t): return 'Factura simplificada'
    if re.search(r'ticket', t): return 'Ticket'
    if re.search(r'nota\s*de\s*cargo', t): return 'Nota de cargo'
    if re.search(r'recibo', t): return 'Recibo'
    return 'Factura'

def detect_cp_pais(text: str) -> str:
    foreign = re.search(
        r'\b(france|germany|italy|portugal|uk|united kingdom|netherlands|belgium|'
        r'luxembourg|usa|united states|ireland|austria|sweden|denmark|norway|'
        r'switzerland|polska|romania|czech|slovakia)\b', text, re.IGNORECASE)
    if foreign:
        return foreign.group(1).title()
    cp_m = re.search(r'\b(\d{5})\b', text)
    return cp_m.group(1) if cp_m else '-'

# ── PDF parser ─────────────────────────────────────────────────────────────────

def extract_text_pdfplumber(file_bytes: bytes) -> str:
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            return '\n'.join(p.extract_text() or '' for p in pdf.pages)
    except Exception:
        return ''

def extract_text_ocr(file_bytes: bytes) -> str:
    try:
        import fitz
        import pytesseract
        from PIL import Image
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        texts = []
        for page in doc:
            pix = page.get_pixmap(dpi=200)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            texts.append(pytesseract.image_to_string(img, lang='spa+eng'))
        return '\n'.join(texts)
    except ImportError:
        return '__OCR_NOT_INSTALLED__'
    except Exception:
        return ''

def parse_pdf(file_bytes: bytes, filename: str, use_ocr: bool) -> dict:
    row = _empty_row(filename)
    ocr_warning = False
    text = extract_text_pdfplumber(file_bytes)
    if len(text.strip()) < 200 and use_ocr:
        ocr_text = extract_text_ocr(file_bytes)
        if ocr_text == '__OCR_NOT_INSTALLED__':
            ocr_warning = True
        elif ocr_text:
            text = ocr_text
    if len(text.strip()) < 10:
        row['error'] = 'PDF sin texto legible.'
        return row
    tl = text.lower()
    row['tipo_documento'] = detect_tipo_doc(text)
    raw_fecha = extract_pat(tl, 'fecha') or extract_pat(text, 'fecha')
    row['fecha'] = parse_date(raw_fecha)
    row['numero_factura'] = extract_pat(text, 'numero') or '-'
    row['nif'] = extract_pat(text, 'nif') or '-'
    nombre = extract_pat(text, 'nombre')
    if not nombre:
        for line in text.split('\n')[:6]:
            line = line.strip()
            if len(line) > 4 and line == line.upper() and not re.match(r'^\d', line):
                nombre = line.title()
                break
    row['nombre_proveedor'] = nombre or '-'
    row['concepto'] = guess_concepto(text)
    dir_raw = extract_pat(text, 'direccion')
    row['direccion'] = dir_raw[:70] if dir_raw else '-'
    row['cp_pais'] = detect_cp_pais(text)
    row['base_imponible'] = parse_amount(extract_pat(tl, 'base') or extract_pat(text, 'base'))
    row['tipo_iva'] = parse_amount(extract_pat(tl, 'pct_iva'))
    row['porcion_iva'] = parse_amount(extract_pat(tl, 'cuota_iva') or extract_pat(text, 'cuota_iva'))
    row['tipo_retencion'] = parse_amount(extract_pat(tl, 'pct_ret'))
    row['porcion_retencion'] = parse_amount(extract_pat(tl, 'imp_ret') or extract_pat(text, 'imp_ret'))
    row['total'] = parse_amount(extract_pat(tl, 'total') or extract_pat(text, 'total'))
    if ocr_warning:
        row['error'] = 'OCR no instalado; resultado sin OCR.'
    row['validacion'] = validate_invoice(row)
    return row

# ── XML Facturae parser ────────────────────────────────────────────────────────

FACTURAE_NS = [
    'http://www.facturae.gob.es/formato/Versiones/Facturaev3_2_2.xsd',
    'http://www.facturae.gob.es/formato/Versiones/Facturaev3_2_1.xsd',
    'http://www.facturae.gob.es/formato/Versiones/Facturaev3_2.xsd',
    '',
]

def _xml_find(root, *paths) -> str:
    for ns in FACTURAE_NS:
        prefix = '{' + ns + '}' if ns else ''
        for path in paths:
            ns_path = '/'.join(prefix + seg for seg in path.split('/'))
            el = root.find('.//' + ns_path)
            if el is not None and el.text:
                return el.text.strip()
    return ''

def parse_xml(file_bytes: bytes, filename: str) -> dict:
    row = _empty_row(filename)
    try:
        root = ET.fromstring(file_bytes)
    except ET.ParseError as e:
        row['error'] = f'XML inválido: {e}'
        return row
    row['tipo_documento'] = 'Factura'
    row['fecha'] = parse_date(_xml_find(root, 'IssueDate'))
    serie  = _xml_find(root, 'InvoiceSeriesCode')
    numero = _xml_find(root, 'InvoiceNumber')
    row['numero_factura'] = f"{serie}-{numero}" if serie and numero else (numero or serie or '-')
    nif  = _xml_find(root, 'SellerParty/TaxIdentification/TaxIdentificationNumber')
    name = (_xml_find(root, 'SellerParty/LegalEntity/CorporateName') or
            _xml_find(root, 'SellerParty/Individual/Name'))
    row['nif'] = nif or '-'
    row['nombre_proveedor'] = name or '-'
    addr = _xml_find(root, 'SellerParty/LegalEntity/AddressInSpain/Address')
    row['direccion'] = addr[:70] if addr else '-'
    cp = _xml_find(root, 'SellerParty/LegalEntity/AddressInSpain/PostCode')
    row['cp_pais'] = cp or '-'
    row['concepto'] = guess_concepto((name or '') + ' ' + (addr or ''))
    row['base_imponible'] = parse_amount(_xml_find(root, 'InvoiceTotals/TotalTaxableBase', 'TaxableBaseAmount'))
    row['tipo_iva']    = parse_amount(_xml_find(root, 'Tax/TaxRate'))
    row['porcion_iva'] = parse_amount(_xml_find(root, 'Tax/TaxAmount/TotalAmount', 'TaxAmount'))
    row['tipo_retencion']    = parse_amount(_xml_find(root, 'WithholdingTax/WithholdingTaxRate'))
    row['porcion_retencion'] = parse_amount(_xml_find(root, 'WithholdingTax/WithholdingTaxAmount/TotalAmount'))
    row['total'] = parse_amount(_xml_find(root, 'InvoiceTotals/TotalGrossAmount', 'InvoiceTotals/TotalInvoiceAmount'))
    row['validacion'] = validate_invoice(row)
    return row

def extract_iva_detail_xml(file_bytes: bytes, filename: str) -> list:
    details = []
    try:
        root = ET.fromstring(file_bytes)
    except ET.ParseError:
        return details
    for ns in FACTURAE_NS:
        prefix = '{' + ns + '}' if ns else ''
        taxes = root.findall('.//' + prefix + 'Tax')
        if taxes:
            for tax in taxes:
                def tf(tag):
                    el = tax.find('.//' + prefix + tag)
                    return el.text.strip() if el is not None and el.text else ''
                details.append({
                    'archivo': filename,
                    'iva_porcentaje': parse_amount(tf('TaxRate')),
                    'base':  parse_amount(tf('TaxableBase') or tf('TaxableBaseAmount')),
                    'cuota': parse_amount(tf('TaxAmount') or tf('TotalAmount')),
                    'raw':   f"Rate={tf('TaxRate')} Base={tf('TaxableBase')} Cuota={tf('TaxAmount')}",
                })
            break
    return details

# ── Exportación Excel ──────────────────────────────────────────────────────────

def to_excel(df: pd.DataFrame, iva_details: list) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        df_out = df.rename(columns=DISPLAY_NAMES)
        df_out.to_excel(writer, sheet_name='Facturas', index=False)
        ws = writer.sheets['Facturas']
        header_fill = PatternFill('solid', fgColor='2D6A4F')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        fill_alt = PatternFill('solid', fgColor='F0FAF5')
        thin = Side(style='thin', color='D0D0D0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for i, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if i % 2 == 0:
                    cell.fill = fill_alt
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 45)
        ws.row_dimensions[1].height = 22
        if iva_details:
            pd.DataFrame(iva_details).to_excel(writer, sheet_name='DetalleIVA', index=False)
    return buf.getvalue()

# ── Exportación PDF ────────────────────────────────────────────────────────────

PDF_COLS    = ['tipo_documento','fecha','numero_factura','nombre_proveedor','nif',
               'concepto','cp_pais','base_imponible','tipo_iva','porcion_iva',
               'tipo_retencion','porcion_retencion','total','validacion']
PDF_HEADERS = ['Tipo','Fecha','Nº Factura','Proveedor','NIF',
               'Concepto','CP/País','Base €','IVA%','Cuota IVA €',
               'Ret%','Ret €','Total €','Valid.']

def _fmt(val) -> str:
    if val is None or val == '-' or (isinstance(val, float) and pd.isna(val)):
        return '-'
    if isinstance(val, float):
        return f"{val:,.2f}".replace(',','X').replace('.',',').replace('X','.')
    s = str(val)
    return (s[:28] + '…') if len(s) > 28 else s

def to_pdf_summary(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
        leftMargin=1.2*cm, rightMargin=1.2*cm, topMargin=1.8*cm, bottomMargin=1.8*cm)
    title_style = ParagraphStyle('T', fontName='Helvetica-Bold', fontSize=14,
        textColor=colors.HexColor('#1C1917'), spaceAfter=2, alignment=TA_CENTER)
    sub_style = ParagraphStyle('S', fontName='Helvetica', fontSize=8,
        textColor=colors.HexColor('#6B6460'), spaceAfter=12, alignment=TA_CENTER)
    foot_style = ParagraphStyle('F', fontName='Helvetica', fontSize=8,
        textColor=colors.HexColor('#2D6A4F'))
    elements = [
        Paragraph("Resumen de Facturas", title_style),
        Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}", sub_style),
        Spacer(1, 0.3*cm),
    ]
    table_data = [PDF_HEADERS]
    for _, row in df.iterrows():
        table_data.append([_fmt(row.get(c)) for c in PDF_COLS])
    col_widths = [2.2*cm,2.0*cm,2.5*cm,3.8*cm,2.2*cm,
                  3.0*cm,1.8*cm,2.0*cm,1.2*cm,2.0*cm,
                  1.2*cm,1.8*cm,2.0*cm,1.4*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#2D6A4F')),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 7),
        ('FONTNAME',    (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',    (0,1), (-1,-1), 6.5),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F0FAF5')]),
        ('GRID',        (0,0), (-1,-1), 0.3, colors.HexColor('#D1D9D1')),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',       (7,1), (12,-1), 'RIGHT'),
        ('TOPPADDING',  (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0),(-1,-1), 3),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING',(0,0), (-1,-1), 4),
    ]))
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        if 'Revisar' in str(row.get('validacion', '')):
            t.setStyle(TableStyle([('BACKGROUND',(13,i),(13,i), colors.HexColor('#FEF3C7'))]))
        if row.get('error'):
            t.setStyle(TableStyle([('BACKGROUND',(0,i),(-1,i), colors.HexColor('#FEF9F9'))]))
    elements.append(t)
    elements.append(Spacer(1, 0.5*cm))
    try:
        tb = df['base_imponible'].apply(safe_float).sum()
        ti = df['porcion_iva'].apply(safe_float).sum()
        tr = df['porcion_retencion'].apply(safe_float).sum()
        tt = df['total'].apply(safe_float).sum()
        line = (f"TOTALES  ·  Base: {_fmt(tb)} €   IVA: {_fmt(ti)} €   "
                f"Retención: {_fmt(tr)} €   Total: {_fmt(tt)} €")
        elements.append(Paragraph(line, foot_style))
    except Exception:
        pass
    doc.build(elements)
    return buf.getvalue()

# ── UI ─────────────────────────────────────────────────────────────────────────

def render_metric(label, value, neutral=False):
    cls = 'neutral' if neutral else ''
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-label">{label}</div>
        <div class="metric-value {cls}">{value}</div>
    </div>""", unsafe_allow_html=True)

def main():
    # Sidebar
    with st.sidebar:
        st.markdown('<div class="sidebar-logo">🧾 FacturasAI</div>', unsafe_allow_html=True)
        st.markdown('<div class="sidebar-tagline">Extracción automática de datos contables para autónomos españoles.</div>', unsafe_allow_html=True)
        st.divider()
        st.markdown('<div class="sidebar-label">Opciones</div>', unsafe_allow_html=True)
        use_ocr = st.checkbox("Activar OCR (PDFs escaneados)", value=False,
            help="Requiere pytesseract, PyMuPDF y Tesseract instalados en el sistema.")
        st.divider()
        st.markdown('<div class="sidebar-label">Columnas extraídas</div>', unsafe_allow_html=True)
        for c in ["Tipo documento","Fecha","Nº Factura","Nombre proveedor","NIF/CIF",
                  "Concepto (categoría)","Dirección","CP / País",
                  "Base imponible (Mod.303)","Tipo IVA %","Porción IVA €",
                  "Tipo Retención %","Porción Retención €","Total €"]:
            st.markdown(f"<span style='font-size:.78rem;color:#6B6460;'>· {c}</span>", unsafe_allow_html=True)
        st.divider()
        st.markdown("<span style='font-size:.73rem;color:#94A3B8;'>v2.0 · Uso privado · Sin almacenamiento de datos</span>", unsafe_allow_html=True)

    # Cabecera
    st.markdown('<p class="main-title">Gestor de Facturas · España</p>', unsafe_allow_html=True)
    st.markdown('<p class="main-subtitle">Sube tus facturas en PDF o XML Facturae y obtén los datos contables organizados en segundos.</p>', unsafe_allow_html=True)
    st.markdown('<div class="alert-priv">🔒 <strong>Privacidad:</strong> Las facturas contienen datos fiscales sensibles. Usa esta herramienta en un entorno privado y de confianza.</div>', unsafe_allow_html=True)

    # Upload
    st.markdown('<div class="section-card"><div class="section-title">📂 Subir facturas</div>', unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        "Arrastra aquí tus archivos",
        type=['pdf','xml'], accept_multiple_files=True, label_visibility='collapsed')
    st.markdown('</div>', unsafe_allow_html=True)

    if not uploaded_files:
        st.markdown('<div style="text-align:center;padding:2.5rem 0;color:#9CA3AF;font-size:.9rem;">⬆️ &nbsp;Sube al menos un archivo <strong>.pdf</strong> o <strong>.xml</strong> (Facturae) para empezar.</div>', unsafe_allow_html=True)
        return

    col_btn, col_info = st.columns([1,3])
    with col_btn:
        process = st.button("▶ Procesar facturas", type="primary", use_container_width=True)
    with col_info:
        st.markdown(f"<span style='font-size:.85rem;color:#6B6460;line-height:2.6;'>{len(uploaded_files)} archivo(s) seleccionado(s)</span>", unsafe_allow_html=True)

    if process:
        rows, iva_details = [], []
        prog = st.progress(0, text="Analizando facturas…")
        n = len(uploaded_files)
        for i, f in enumerate(uploaded_files):
            prog.progress((i+1)/n, text=f"Procesando {f.name}…")
            data = f.read()
            ext = f.name.rsplit('.',1)[-1].lower()
            if ext == 'pdf':
                row = parse_pdf(data, f.name, use_ocr)
            elif ext == 'xml':
                row = parse_xml(data, f.name)
                iva_details.extend(extract_iva_detail_xml(data, f.name))
            else:
                row = _empty_row(f.name)
                row['error'] = 'Formato no soportado.'
            rows.append(row)
        prog.empty()
        df = pd.DataFrame(rows, columns=COLUMNS)
        st.session_state['df'] = df
        st.session_state['iva_details'] = iva_details
        st.markdown(f'<div class="alert-ok">✓ {n} archivo(s) procesado(s) correctamente.</div>', unsafe_allow_html=True)

    if 'df' not in st.session_state:
        return

    df = st.session_state['df']
    iva_details = st.session_state.get('iva_details', [])

    # Métricas
    st.markdown("<br>", unsafe_allow_html=True)
    mc = st.columns(4)
    with mc[0]: render_metric("Facturas procesadas", str(len(df)), neutral=True)
    with mc[1]: render_metric("Base total", fmt_eur(df['base_imponible'].apply(safe_float).sum()))
    with mc[2]: render_metric("IVA total", fmt_eur(df['porcion_iva'].apply(safe_float).sum()))
    with mc[3]: render_metric("Total", fmt_eur(df['total'].apply(safe_float).sum()))
    st.markdown("<br>", unsafe_allow_html=True)

    # Alertas
    revisiones = df[df['validacion'].astype(str).str.contains('Revisar', na=False)]
    errores    = df[df['error'].astype(str).str.len() > 0]
    if not revisiones.empty:
        st.markdown(f'<div class="alert-warn">⚠️ {len(revisiones)} factura(s) con importes que no cuadran — revisa la columna <strong>Validación</strong>.</div>', unsafe_allow_html=True)
    if not errores.empty:
        st.markdown(f'<div class="alert-err">❌ {len(errores)} archivo(s) con errores de parseo.</div>', unsafe_allow_html=True)

    # Tabla editable
    st.markdown('<div class="section-card"><div class="section-title">📋 Datos extraídos — edita si es necesario</div>', unsafe_allow_html=True)
    edited_df = st.data_editor(
        df,
        column_config={
            'tipo_documento':   st.column_config.SelectboxColumn('Tipo', options=TIPO_DOC_OPTIONS, width='small'),
            'fecha':            st.column_config.TextColumn('Fecha', width='small'),
            'numero_factura':   st.column_config.TextColumn('Nº Factura', width='small'),
            'nombre_proveedor': st.column_config.TextColumn('Proveedor', width='medium'),
            'nif':              st.column_config.TextColumn('NIF/CIF', width='small'),
            'concepto':         st.column_config.SelectboxColumn('Concepto', options=CONCEPTO_OPTIONS, width='medium'),
            'direccion':        st.column_config.TextColumn('Dirección', width='medium'),
            'cp_pais':          st.column_config.TextColumn('CP / País', width='small'),
            'base_imponible':   st.column_config.NumberColumn('Base (€)', format="%.2f", width='small'),
            'tipo_iva':         st.column_config.NumberColumn('IVA %', format="%.0f", width='small'),
            'porcion_iva':      st.column_config.NumberColumn('Cuota IVA (€)', format="%.2f", width='small'),
            'tipo_retencion':   st.column_config.NumberColumn('Ret. %', format="%.0f", width='small'),
            'porcion_retencion':st.column_config.NumberColumn('Ret. (€)', format="%.2f", width='small'),
            'total':            st.column_config.NumberColumn('Total (€)', format="%.2f", width='small'),
            'validacion':       st.column_config.TextColumn('Valid.', disabled=True, width='small'),
            'error':            st.column_config.TextColumn('Error', disabled=True, width='medium'),
        },
        use_container_width=True, num_rows="dynamic", key='editor', hide_index=True,
    )
    st.markdown('</div>', unsafe_allow_html=True)

    for idx, row in edited_df.iterrows():
        edited_df.at[idx, 'validacion'] = validate_invoice(row.to_dict())

    if iva_details:
        with st.expander(f"🔍 Detalle de tramos IVA — {len(iva_details)} registro(s)"):
            st.dataframe(pd.DataFrame(iva_details), use_container_width=True, hide_index=True)

    # Exportar
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-card"><div class="section-title">⬇️ Exportar resultados</div>', unsafe_allow_html=True)
    ca, cb = st.columns(2)
    with ca:
        try:
            xls = to_excel(edited_df, iva_details)
            st.download_button("📊  Descargar Excel (.xlsx)", data=xls,
                file_name=f"facturas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True)
            st.markdown("<span style='font-size:.76rem;color:#9CA3AF;'>Hoja Facturas + DetalleIVA si hay múltiples tramos</span>", unsafe_allow_html=True)
        except Exception as e:
            st.error(f"Error Excel: {e}")
    with cb:
        try:
            pdf_b = to_pdf_summary(edited_df)
            st.download_button("📄  Descargar PDF resumen", data=pdf_b,
                file_name=f"resumen_facturas_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime='application/pdf', use_container_width=True)
            st.markdown("<span style='font-size:.76rem;color:#9CA3AF;'>Tabla consolidada · A4 horizontal · con totales</span>", unsafe_allow_html=True)
        except Exception as e:
            st.error(f"Error PDF: {e}")
    st.markdown('</div>', unsafe_allow_html=True)

if __name__ == '__main__':
    main()
